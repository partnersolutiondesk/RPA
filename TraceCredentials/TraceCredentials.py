from openpyxl import Workbook
import json
import traceback
import pandas as pd
import shutil
import inspect
import logging
import csv
import datetime as dt
from pathlib import Path
import requests
from requests_toolbelt.multipart.encoder import MultipartEncoder
from openpyxl.styles import Font
# Global variables defined
logger = logging.getLogger(__name__)
updates_json = None


# Log the messages using logging library
def log(log_msg, log_level):
    # Automatically log the current function details.
    # Get the previous frame in the stack, otherwise it would be this function!!!
    func = inspect.currentframe().f_back.f_code

    # Dump the message + the name of this function to the log.
    logger.log(level=getattr(logging, log_level.upper(), None), msg='{0}): {1}'.format(func.co_name, log_msg))


# Initializing a logger with custom configuration
def initialize_logger(log_file_path, log_level):
    logger.setLevel(getattr(logging, log_level.upper()))
    file_handler = logging.FileHandler(log_file_path, mode='a')
    formatter = logging.Formatter('%(asctime)s %(levelname)s %(name)s (%(message)s', datefmt='(%d-%m-%Y %I.%M.%S %p)')

    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    log("Log file started.", 'info')


def get_report(args):
    folder_id = args[0]
    url = args[1]
    username = args[2]
    apikey = args[3]
    inputoutputFolderPath = args[4]

    log_file_path = inputoutputFolderPath + "\\packagescanner.log"

    log_level = 'debug'
    initialize_logger(log_file_path=log_file_path, log_level=log_level)
    log("get_report fun called with the: folder_id: " + folder_id,log_level)
    # Create sheet in input xlsx
    return get_files_from_folder(folder_id, url, username, apikey, inputoutputFolderPath)

def find_matching_nodes(content, match_rules):
    matched = []

    nodes = content.get("nodes", [])
    variables = content.get("variables", [])

    # Normalize match rules (support typo)
    rule_names = {
        str(
            rule.get("credentialName")
        ).strip().lower()
        for rule in match_rules
        if rule.get("credentialName")
    }

    # ---- Scan node credentials ----
    for node in nodes:
        for attr in node.get("attributes", []):
            value = attr.get("value", {})
            if value.get("type") == "CREDENTIAL":
                cred = value.get("credential", {})
                cred_name = cred.get("name")

                if cred_name and cred_name.strip().lower() in rule_names:
                    matched.append({
                        "source": "node",
                        "nodeUid": node.get("uid"),
                        "commandName": node.get("commandName"),
                        "credentialName": cred_name,
                        "lockerName": cred.get("lockerName"),
                        "attributeName": cred.get("attributeName"),
                    })

    # ---- Scan variable credentials ----
    for var in variables:
        default = var.get("defaultValue", {})
        if default.get("type") == "CREDENTIAL":
            cred = default.get("credential", {})
            cred_name = cred.get("name")

            if cred_name and cred_name.strip().lower() in rule_names:
                matched.append({
                    "source": "variable",
                    "variableName": var.get("name"),
                    "credentialName": cred_name,
                    "lockerName": cred.get("lockerName"),
                    "attributeName": cred.get("attributeName"),
                })

    return matched





import openpyxl

def load_input_match_rules(filepath):
    wb = openpyxl.load_workbook(filepath)

    if "Inputs" not in wb.sheetnames:
        raise ValueError(f'Sheet "Inputs" not found in {filepath}')

    sheet = wb["Inputs"]
    match_rules = []

    # ✅ Expecting columns: PackageName (instead of CommandName)
    for row in sheet.iter_rows(min_row=1, values_only=True):
        credential_name = str(row[0]).strip() if row[0] is not None else ""
        # version = str(row[1]).strip() if row[1] is not None else ""

        if credential_name :
            # Only package name is present
            match_rules.append({
                "credentialName": credential_name
            })


    log(', '.join(str(i) for i in match_rules), "debug")
    return match_rules


def get_file_content(file_id, token, url, username, apikey):
    try:
        headers = {
            "X-Authorization": token
        }
        log(f"calling content url","debug")
        response = requests.get(f"{url}/v2/repository/files/{file_id}/content", headers=headers)
        if response.status_code == 401:
            log(f"401 Unauthorized error. Re-authenticating for new token","debug")
            token = authenticate(url, username, apikey)
            if not token:
                log("Re-authentication failed.","debug")
                return
            log(f"New token received: {token}","debug")
            headers["X-Authorization"] = str(token)
            log(f"content","debug")
            response = requests.get(
                f"content api call",
                headers=headers
            )
        response.raise_for_status()
        return response.text  # or response.json(), depending on content type
    except requests.RequestException as e:
        log(f"Failed to fetch content for file ID {file_id}: {e}","debug")
        return None


# Auth function to get token
def authenticate(url, username, apikey):
    auth_payload = {
        "username": username,
        "apiKey": apikey
    }
    try:
        response = requests.post(f"{url}/v2/authentication", json=auth_payload)
        response.raise_for_status()
        return response.json().get("token")
    except requests.RequestException as e:
        log(f"Authentication failed: {e}","debug")
        return None



addRowsToExcelCounter = 0


def add_rows_to_excel(filepath, rows, botID):
    """
    Add rows to an Excel sheet. Create the sheet if it doesn't exist.

    Args:
        filepath (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to write to.
        rows (list of list): List of rows to add (each row is a list of cell values).
    """
    global addRowsToExcelCounter
    addRowsToExcelCounter += 1
    log("Going to add rows of " + botID, "debug")
    header = ["Bot Id", "Bot Name", "Credential Name"]
    try:
        wb = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        wb = Workbook()

    # Create sheet if it doesn't exist
    if "Report" in wb.sheetnames:
        sheet = wb["Report"]
    else:
        sheet = wb.create_sheet(title="Report")
        sheet.append(header)

        # Append rows
    for row in rows:
        sheet.append(row)

    # Update BotIDs sheet
    headerBotID = ["Bot Id"]
    if "Bot_IDs" in wb.sheetnames:
        sheet2 = wb["Bot_IDs"]
    else:
        sheet2 = wb.create_sheet(title="Bot_IDs")
        sheet2.append(headerBotID)
    sheet2.append([botID])
    log("Added bot ID: " + botID + " specific rows to the input excel's Report sheet" , "debug")
    wb.save(filepath)


def add_empty_data_message_to_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.create_sheet(title="Report")
        # row = ["No data found for enabled bot commands requiring updates as per the input commands"]
        # sheet.append(row)

        message_lines = [
            "NO MATCHING COMMANDS WERE FOUND. Possible reasons include:",
            "",
            "1. Only ENABLED command lines are being considered.",
            "2. Commands that already have the expected value from the input are excluded."
        ]

        # Styles
        bold_red = Font(bold=True, color="FF0000")  # Red
        bold_black = Font(bold=True, color="000000")  # Black

        # Write each line into the Excel sheet
        for i, line in enumerate(message_lines, start=1):
            cell = sheet.cell(row=i, column=1, value=line)
            cell.font = bold_red if i == 1 else bold_black
        wb.save(filepath)
        log("Added new row to the input excel's Report sheet: " , "debug")
    except FileNotFoundError:
        wb = Workbook()


# Main function to get files
def get_files_from_folder(folder_id, url, username, apikey, inputoutputFolderPath):
    global addRowsToExcelCounter
    token = authenticate(url, username, apikey)
    if not token:
        log("Failed to authenticate. Exiting.","debug")
        return None

    all_files = []
    botIDs = []
    seen_items = set()  # Track IDs to prevent duplicates
    processed_subfolders = set()  # Track processed subfolders
    matched_files = []
    match_rules = load_input_match_rules(inputoutputFolderPath + "\\input.xlsx")

    def fetch_folder(folder_id, token):
        log("inside fetch folder" + folder_id, "debug")
        nonlocal all_files
        page_number = 0

        while True:
            headers = {
                "X-Authorization": str(token)
            }

            payload = {
                "filter": {},
                "sort": [{"field": "id", "direction": "desc"}],
                "page": {
                    "offset": page_number * 100,
                    "length": 100
                }
            }
            log(f"calling folder API with token {token}","debug")
            response = requests.post(
                f"{url}/v2/repository/folders/{folder_id}/list",
                headers=headers,
                json=payload
            )

            if response.status_code == 401:
                log("401 unauthorized error. Re-authenticating...","debug")
                token = authenticate(url, username, apikey)

                if not token:
                    log("Re-authentication failed.","debug")
                    return
                log(f"New token generated {token}","debug" )
                headers["X-Authorization"] = str(token)
                log(f"Calling /v2/repository/folders/{folder_id}/list API with token {token}","debug")
                response = requests.post(
                    f"{url}/v2/repository/folders/{folder_id}/list",
                    headers=headers,
                    json=payload
                )
            print(response.raise_for_status())
            data = response.json()
            response.raise_for_status()
            items = data.get("list", [])

            # Stop pagination if no new items are received
            new_items = [item for item in items if item['id'] not in seen_items]
            if not new_items:
                break

            for item in new_items:
                seen_items.add(item['id'])  # Track item ID to avoid duplication

                if item['type'] == 'application/vnd.aa.taskbot':
                    matched_files = []
                    log("Found a taskbot" + item['id'], "debug")
                    all_files.append(item)  # Store full item details
                    content = get_file_content(item['id'], token, url, username, apikey)
                    print("Content is",content)
                    if not content:
                        continue

                    try:
                        content_json = json.loads(content)
                        # print(content_json)
                        log("Analysing taskbot content", "debug")

                        all_matches = find_matching_nodes(content_json, match_rules)  # ✅ pass entire content, not each node
                        for match in all_matches:
                            log("Match nodes found", "debug")
                            matched_files.append([
                                item['id'],
                                item['name'],
                                # item['path'],
                                match.get("credentialName")
                                # match.get("packageName")
                            ])
                        # insert each bot specific rows into excel in one go
                        if len(matched_files) > 0:
                            log("matched_files length more than 0. so adding it to the input excel Report sheet", "debug")
                            add_rows_to_excel(
                                inputoutputFolderPath + "\\input.xlsx",
                                matched_files,
                                item['id']
                            )




                    except Exception as e:
                        log(f"JSON parse error for {item.get('name')}: {e}","debug")

                elif item['type'] == 'application/vnd.aa.directory' and item['id'] not in processed_subfolders:
                    processed_subfolders.add(item['id'])  # Mark subfolder as processed
                    fetch_folder(item['id'], token)  # Recursively fetch subfolder contents

            page_number += 1  # Move to the next page for pagination

    fetch_folder(folder_id, token)
    formatted_json = json.dumps({"list": all_files}, indent=4)
    if addRowsToExcelCounter == 0:
        add_empty_data_message_to_excel(inputoutputFolderPath+"\\input.xlsx")
    # print(formatted_json)
    return formatted_json


args = ['folder-id','cr-url','syed.hasnain','api-key',r'folder-path']
get_report(args)