

import os, time, logging, requests, openpyxl
from collections import defaultdict
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────
CR_URL      = os.getenv("AA_CR_URL",      "CR_URL")
USERNAME    = os.getenv("AA_USERNAME",    "CR_username")
PASSWORD    = os.getenv("AA_PASSWORD",    "CR_password")
ROOT_FOLDER = os.getenv("AA_FOLDER_ID",  "CR_folder_id")         # top-level folder ID to scan
PAGE_SIZE   = 400                                        # items per page (max 100)
DELAY       = 0.15                                       # seconds between API calls
VERIFY_SSL  = True    # set False for self-signed certs
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

DEPRECATED_APIS = [
    "v1/schedule/automations ",
    "v1/authentication/token",
    "v1/authentication",
    "v3/wlm/workitemmodel",
    "v3/wlm/workitemmodels/{id}",
    "v3/wlm/queues/{id}",
    "v3/wlm/queues/{id}/consumers",
    "v3/wlm/queues/{id}/participants",
    "v3/wlm/queues/{id}/members/{userId}",
    "v3/wlm/queues/{id}/file",
    "v3/wlm/queues/{id}/workitems/{workItemId}"
]

# ─────────────────────────────────────────────────────
#  API CALLS
# ─────────────────────────────────────────────────────

def authenticate() -> requests.Session:
    """POST /v2/authentication — returns an authorised session."""
    session = requests.Session()
    session.verify = VERIFY_SSL
    resp = session.post(
        f"{CR_URL}/v2/authentication",
        json={"username": USERNAME, "password": PASSWORD},
        headers={"Content-Type": "application/json"},
    )
    resp.raise_for_status()
    token = resp.json().get("token")
    if not token:
        raise ValueError(f"No token in response: {resp.text}")
    session.headers.update({"X-Authorization": token, "Content-Type": "application/json"})
    log.info("Authenticated successfully.")
    return session


def list_folder_contents(session: requests.Session, folder_id: str) -> list:
    """
    POST /v2/repository/folders/{folderId}/list
    Paginates automatically. Returns all items (files + sub-folders).
    """
    url = f"{CR_URL}/v2/repository/folders/{folder_id}/list"
    items, offset = [], 0
    while True:
        body = {
            "sort": [{"field": "name", "direction": "asc"}],
            "page": {"offset": offset, "length": PAGE_SIZE},
        }
        resp = session.post(url, json=body)
        resp.raise_for_status()
        data  = resp.json()
        batch = data.get("list", [])
        items.extend(batch)
        total = data.get("page", {}).get("total", 0)
        offset += len(batch)
        if offset >= total or not batch:
            break
        time.sleep(DELAY)
    return items

# Set to True on first run to dump one bot's raw content JSON for inspection
DEBUG_DUMP_CONTENT = False

# def get_content(session: requests.Session, file_id: str) -> dict:
#     """
#     GET /v2/repository/files/{fileId}/content
#     Returns the full bot definition JSON.
#     Contains both package list and all action nodes (including TaskBot Run → child refs).
#     """
#     resp = session.get(f"{CR_URL}/v2/repository/files/{file_id}/content")
#     if resp.status_code in (400, 404, 403):
#         return {}
#     resp.raise_for_status()
#     content_type = resp.headers.get("Content-Type", "").lower()
#
#     if "application/json" in content_type:
#         data = resp.json()
#     else:
#         # If it's XML or plain text, capture it as a raw string
#         data = resp.text
#
#     # Handle debugging/dumping to files
#     if DEBUG_DUMP_CONTENT:
#         import json as _json
#         dump_path = f"content_debug_{file_id}"
#
#         if isinstance(data, dict):
#             # Save JSON formatted cleanly
#             with open(f"{dump_path}.json", "w", encoding="utf-8") as f:
#                 _json.dump(data, f, indent=2)
#             log.info(f"     📄 Raw JSON content dumped → {dump_path}.json")
#         else:
#             # Save XML or plain text as-is
#             # (Checks if it starts with XML declaration to guess extension)
#             ext = "xml" if data.strip().startswith("<") else "txt"
#             with open(f"{dump_path}.{ext}", "w", encoding="utf-8") as f:
#                 f.write(data)
#             log.info(f"     📄 Raw {ext.upper()} content dumped → {dump_path}.{ext}")
#
#     return data

def get_content(session: requests.Session, file_id: str):
    resp = session.get(f"{CR_URL}/v2/repository/files/{file_id}/content")
    print(f"{CR_URL}/v2/repository/files/{file_id}/content")

    # Check if it failed
    if resp.status_code in (400, 403, 404):
        print(f"❌ API Error {resp.status_code} for file ID {file_id}!")
        print(f"   Server Response: {resp.text}") # <--- This will tell us the exact cause!
        return None

    resp.raise_for_status()

    try:
        return resp.json()
    except ValueError:
        return resp.text

def scan_and_inspect_folders(session: requests.Session, current_folder_id: str, flagged_items: list):
    """
    Recursively steps through folders and safely inspects valid file types.
    """
    print(f"🔍 Scanning folder ID: {current_folder_id}...")
    try:
        items = list_folder_contents(session, current_folder_id)
    except Exception as e:
        log.error(f"Failed to list contents for folder {current_folder_id}: {e}")
        return

    for item in items:
        item_id = item.get("id") or item.get("fileId") or item.get("folderId")
        item_name = item.get("name", "Unknown")
        is_folder = item.get("typeLabel") == "FOLDER"

        if is_folder:
            scan_and_inspect_folders(session, item_id, flagged_items)
        else:
            # 🛑 CRITICAL SAFETY CHECK: Skip binary files that cause 500 errors
            # Add any other extensions your directory has (e.g., .pdf, .zip, .png)
            ignored_extensions = ('.xlsx', '.xls', '.pdf', '.png', '.jpg', '.zip', '.exe', '.doc', '.docx')
            if item_name.lower().endswith(ignored_extensions):
                print(f"   ⏩ Skipping binary asset file: '{item_name}'")
                continue

            # Wrap get_content in a local try/except block so one bad file
            # won't kill your entire script run.
            try:
                content = get_content(session, item_id)
                if content is None:
                    continue
            except requests.exceptions.HTTPError as http_err:
                # Catch the 500 error cleanly, log it, and keep moving to the next file!
                log.error(f"   ❌ Server failed to process file '{item_name}' (ID: {item_id}). Error: {http_err}")
                continue
            except Exception as e:
                log.error(f"   ❌ Unexpected error reading file '{item_name}': {e}")
                continue

            # Convert to text string for universal keyword scanning
            if isinstance(content, dict):
                import json
                content_str = json.dumps(content).lower()
            else:
                content_str = str(content).lower()

            # Check for any deprecated API matches
            found_deprecations = []
            for api in DEPRECATED_APIS:
                if api.lower() in content_str:
                    found_deprecations.append(api)

            if found_deprecations:
                print(f"   ⚠️ WARNING: Found deprecated APIs in file '{item_name}' ({item_id}) -> {found_deprecations}")
                flagged_items.append({
                    "id": item_id,
                    "name": item_name,
                    "deprecated_apis": ", ".join(found_deprecations)
                })
# --- Excel Report Generator ---
def create_excel_report(flagged_items: list, filename: str = "deprecated_apis_report.xlsx"):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deprecated API Audit"

    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True

    # Header and Styling definitions
    headers = ["File ID", "File Name", "Found Deprecated APIs"]
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Subtle Gray
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Write headers
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    # Write data rows
    for row_idx, item in enumerate(flagged_items, start=2):
        row_data = [item["id"], item["name"], item["deprecated_apis"]]
        ws.append(row_data)

        # Apply row alternate shading (Zebra striping) and borders
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = zebra_fill

    # Autofit column widths cleanly
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Freeze the top header row
    ws.freeze_panes = "A2"

    # Save output file
    wb.save(filename)
    print(f"\n📊 Excel report compiled successfully saved as: {filename}")
# ─────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────
def main():
    try:
        print("🔐 Authenticating session...")
        session = authenticate()

        # This list will hold matches across all folder depths
        flagged_items = []

        print("\n🚀 Commencing deep search across folder structure...")
        scan_and_inspect_folders(session, ROOT_FOLDER, flagged_items)

        print(f"\nScan completed. Found {len(flagged_items)} files containing deprecated APIs.")

        # Export finding list even if empty (will just generate structural headers)
        create_excel_report(flagged_items)

    except Exception as e:
        log.error(f"Execution halted due to an unexpected exception: {e}")

if __name__ == "__main__":
    results = main()

